import openpyxl
import pyinputplus as pyip
import os
import pandas as pd

# Function to get the top songs for a specific year
def get_top_songs_by_year(data, year):
    selected_year_songs = [song for song in data if song['released_year'] == year]
    sorted_songs = sorted(selected_year_songs, key=lambda x: x['streams'], reverse=True)
    top_10_songs = sorted_songs[:10]
    return top_10_songs

# Function to search for a song in the database
def search_song(song_name, data):
    for song in data:
        if str(song['track_name']).lower() == str(song_name).lower():
            return song

# Function to display information about a song
def display_song_info(song_info):
    print("Song name:", song_info['track_name'])
    print("Artist name:", song_info['artist_name'])
    print(f"Released date: {song_info['released_year']}-{song_info['released_month']:02d}-{song_info['released_day']:02d}")

# Function to add a song to the favorite album
def add_to_favorite(song_info, favorite_sheet, favorite_file_path):
    favorite_sheet.append([song_info['track_name'], song_info['artist_name'], f"{song_info['released_year']}-{song_info['released_month']:02d}-{song_info['released_day']:02d}"])
    favorite_workbook.save(favorite_file_path)
    print(f"The song '{song_info['track_name']}' by '{song_info['artist_name']}' has been added to your favorite album.")

# Function to remove a song from the favorite album
def remove_from_favorite(song_name, favorite_sheet, favorite_file_path):
    for index, row in enumerate(favorite_sheet.iter_rows(min_row=2, values_only=True), start=2):
        if str(row[0]).lower() == song_name.lower():
            favorite_sheet.delete_rows(index)
            favorite_workbook.save(favorite_file_path)
            print(f"The song '{row[0]}' has been removed from your favorite album.")
            return

# Function to show genre rankings
def show_genre_ranking(data_file):
    data = pd.read_excel(data_file)
    #list of genres
    print("Available genres:")
    print("1. Rock")
    print("2. Pop")
    print("3. Hip-hop")
    print("4. Country")
    print("5. Soul")
    print("6. Folk")
    print("7. Jazz")
    print("8. Heavy Metal")
    print("9. EDM")

    genre_choice = input("Enter the number for the genre you want to explore: ")
        #mapping for genres
    genre_mapping = {
        "1": "Rock",
        "2": "Pop",
        "3": "Hip-hop",
        "4": "Country",
        "5": "Soul",
        "6": "Folk",
        "7": "Jazz",
        "8": "Heavy Metal",
        "9": "EDM",
    }

    selected_genre = genre_mapping.get(genre_choice)

    if selected_genre:
        filtered_data = filter_genre(data, selected_genre)
        display_top_songs(filtered_data, selected_genre)
    else:
        print("Invalid genre choice. Please select a valid genre.")

# Function to filter data based on genre
def filter_genre(data, genre):
    # Filtering logic based on genre criteria
    if genre == "Rock":
        filtered_data = data[(data['energy_%'] > 70) & (data['valence_%'] > 60)]
    elif genre == "Pop":
        filtered_data = data[(data['danceability_%'] > 60) & (data['energy_%'] > 60)]
    elif genre == "Hip-hop":
        filtered_data = data[(data['danceability_%'] > 70) & (data['speechiness_%'] > 50)]
    elif genre == "Country":
        filtered_data = data[(data['acousticness_%'] > 50) & (data['valence_%'] > 50)]
    elif genre == "Soul":
        filtered_data = data[(data['danceability_%'] > 50) & (data['valence_%'] > 50)]
    elif genre == "Folk":
        filtered_data = data[(data['acousticness_%'] > 50) & (data['instrumentalness_%'] > 50)]
    elif genre == "Jazz":
        filtered_data = data[(data['acousticness_%'] > 50) & (data['valence_%'] > 50)]
    elif genre == "Heavy Metal":
        filtered_data = data[(data['valence_%'] > 50) & (data['energy_%'] > 70)]
    elif genre == "EDM":
        filtered_data = data[(data['instrumentalness_%'] > 50) & (data['energy_%'] > 70)]
    else:
        filtered_data = pd.DataFrame()  # Empty DataFrame for unknown genres
    return filtered_data

# Function to display top songs for a specific genre
def display_top_songs(data, genre):
    # Displaying top songs for a genre
    if not data.empty:
        top_10_songs = data.nlargest(10, 'streams')
        print(f"Top 10 streamed songs in the {genre} genre:")
        print(top_10_songs[['track_name', 'released_year', 'artist(s)_name']])
    else:
        print(f"No data available for the {genre} genre.")
    
# Get the directory of the current script
script_dir = os.path.dirname(os.path.abspath(__file__))

# Create or load the favorite workbook
favorite_file_path = os.path.join(script_dir, 'favorite_songs.xlsx')
if os.path.exists(favorite_file_path):
    favorite_workbook = openpyxl.load_workbook(favorite_file_path)
else:
    favorite_workbook = openpyxl.Workbook()
    favorite_workbook.active.append(["track_name", "artist_name", "Released date"])

favorite_sheet = favorite_workbook.active

# Open the database workbook
file_path = os.path.join(script_dir, 'database.xlsx')
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# Read Excel file
data = []
for row in sheet.iter_rows(min_row=2, values_only=True):
    song = {
        'track_name': row[0],
        'artist_name': row[1],
        'released_year': row[3],
        'released_month': row[4],
        'released_day': row[5],
        'streams': row[8]
    }
    data.append(song)

# Main menu loop
while True:
    print("\nMenu:")
    print("1. Search, Add, or Remove Songs")
    print("2. Display Top 10 Songs by Year")
    print("3. Genre")
    print("4. Exit")

    choice = pyip.inputInt("Enter your choice (1-4): ", min=1, max=4)
        # Code for searching, adding, or removing songs
    if choice == 1:
        user_input = input("Please input the song name: ")
        song_result = search_song(user_input, data)

        if song_result:
            display_song_info(song_result)

            # Ask the user if they want to add, remove, or do nothing
            action_menu = ['Add to favorite', 'Remove from favorite', 'Do nothing', 'Display Top 10 by Year']
            action = pyip.inputMenu(action_menu, numbered=True, prompt='Choose an action: ')

            if action == 'Add to favorite':
                add_to_favorite(song_result, favorite_sheet, favorite_file_path)
            elif action == 'Remove from favorite':
                remove_from_favorite(user_input, favorite_sheet, favorite_file_path)
            elif action == 'Display Top 10 by Year':
                # Allow the user to display top 10 songs for a specific year
                user_year = pyip.inputInt("Enter the year you want to view: ", min=2000, max=9999)
                top_songs = get_top_songs_by_year(data, user_year)

                print(f"\nTop 10 songs of {user_year}:\n")
                for i, song in enumerate(top_songs, start=1):
                    print(f"{i}. {song['track_name']} - {song['artist_name']}, Streams: {song['streams']}")

        else:
            print(f" '{user_input}' not found in the database.")


        # Extract the year the user wants to view
    elif choice == 2:
        # Code for displaying top 10 songs by year
        user_year = pyip.inputInt("Enter the year you want to view: ", min=2000, max=9999)

        try:
            if user_year not in {song['released_year'] for song in data}:
                raise ValueError(f"{user_year} is not in the data.")

            # Get the Top 10 songs
            top_songs = get_top_songs_by_year(data, user_year)

            # Print the results
            print(f"\nTop 10 songs of {user_year}:\n")
            for i, song in enumerate(top_songs, start=1):
                print(f"{i}. {song['track_name']} - {song['artist_name']}, Streams: {song['streams']}")

        except ValueError as ve:
            print(f"Error: {ve}")

    elif choice == 3:
        # Code for genre exploration
        show_genre_ranking('database.xlsx')

    elif choice == 4:
        # Code for exiting the program
        workbook.close()
        favorite_workbook.close()
        break
